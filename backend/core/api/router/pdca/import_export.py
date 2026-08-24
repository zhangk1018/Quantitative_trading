"""
import_export.py - Excel 导入/导出/备份 API
"""
import os
import io
import logging
import subprocess
from typing import Optional
from datetime import datetime

from fastapi import APIRouter, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from core.api.dependencies import get_db
from shared.error_codes import CommonError, PDCAError
from shared.schemas import ApiResponse

logger = logging.getLogger(__name__)
router = APIRouter(tags=["导入导出"])

# 免责声明文本
_DISCLAIMER = "免责声明：本软件仅为交易记录复盘工具，不构成投资建议，不保证盈利"


# ============================================================
# 导出：台账 Excel
# ============================================================

@router.get("/export", response_model=ApiResponse)
async def export_records_redirect(cycle_id: Optional[int] = Query(None, alias="cycle_id")):
    """导出台账 Excel（别名，兼容前端调用）"""
    return await export_records(cycle_id)


@router.get("/export/records", response_model=ApiResponse)
async def export_records(cycle_id: Optional[int] = Query(None, alias="cycle_id")):
    """导出台账 Excel（含免责声明首行）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="50002: 缺少 openpyxl 库，无法导出 Excel")

    with get_db() as conn:
        with conn.cursor() as cur:
            if cycle_id:
                cur.execute(
                    """
                    SELECT * FROM pdca.trading_record
                    WHERE pdca_cycle_id = %s AND deleted_at IS NULL
                    ORDER BY entry_date DESC
                    """,
                    (cycle_id,),
                )
            else:
                cur.execute(
                    "SELECT * FROM pdca.trading_record WHERE deleted_at IS NULL ORDER BY entry_date DESC"
                )

            columns = [desc[0] for desc in cur.description]
            rows = cur.fetchall()

    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交易台账"

    # 免责声明行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    disclaimer_cell = ws.cell(row=1, column=1, value=_DISCLAIMER)
    disclaimer_cell.font = Font(bold=True, color="FF0000")
    disclaimer_cell.alignment = Alignment(horizontal="center")

    # 表头
    header_row = 2
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # 数据
    for row_idx, row in enumerate(rows, header_row + 1):
        for col_idx, value in enumerate(row, 1):
            if isinstance(value, datetime):
                value = value.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 输出到 BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"pdca_trading_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ============================================================
# 导出：复盘报告
# ============================================================

@router.get("/export/report/{report_id}")
async def export_report(report_id: int):
    """导出复盘报告为 Excel（含免责声明）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="50002: 缺少 openpyxl 库")

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM pdca.pdca_check_report WHERE id = %s",
                (report_id,),
            )
            columns = [desc[0] for desc in cur.description]
            row = cur.fetchone()

    if not row:
        raise HTTPException(status_code=404, detail=PDCAError.REPORT_NOT_FOUND.detail())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "复盘报告"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    disclaimer_cell = ws.cell(row=1, column=1, value=_DISCLAIMER)
    disclaimer_cell.font = Font(bold=True, color="FF0000")
    disclaimer_cell.alignment = Alignment(horizontal="center")

    header_row = 2
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    for col_idx, value in enumerate(row, 1):
        if isinstance(value, datetime):
            value = value.isoformat()
        ws.cell(row=header_row + 1, column=col_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"pdca_check_report_{report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ============================================================
# 全量备份
# ============================================================

@router.get("/backup")
async def backup_pdca_get():
    """全量备份 pdca schema（GET 方式，兼容前端调用）"""
    return await backup_pdca_post()


@router.post("/export/backup")
async def backup_pdca_post():
    """全量备份 pdca schema（pg_dump 生成 SQL 文件）"""
    try:
        result = subprocess.run(
            [
                "pg_dump",
                "--host", os.environ.get("PG_HOST", "localhost"),
                "--port", os.environ.get("PG_PORT", "5432"),
                "--username", os.environ.get("PG_USER", "quant_user"),
                "--dbname", os.environ.get("PG_DATABASE", "quant_trading"),
                "--schema", "pdca",
                "--format", "c",
                "--file", "-",
            ],
            capture_output=True,
            text=False,
            env={
                **os.environ,
                "PGPASSWORD": os.environ.get("PG_PASSWORD", ""),
            },
        )
        if result.returncode != 0:
            logger.error("pg_dump 失败: %s", result.stderr.decode())
            raise HTTPException(status_code=500, detail="50002: 备份失败")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            io.BytesIO(result.stdout),
            media_type="application/octet-stream",
            headers={
                "Content-Disposition": f"attachment; filename=pdca_backup_{timestamp}.dump",
            },
        )
    except FileNotFoundError:
        raise HTTPException(status_code=500, detail="50002: pg_dump 未安装")
    except Exception as e:
        logger.exception("备份失败")
        raise HTTPException(status_code=500, detail=f"50002: {str(e)}")


# ============================================================
# 券商导入
# ============================================================

@router.post("/import/parse", response_model=ApiResponse)
async def parse_import(
    file: UploadFile = File(...),
    broker_name: str = Query(..., description="券商名称（如 htsc, citics）"),
):
    """解析券商 Excel（华泰/中信适配器）"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="50002: 缺少 openpyxl 库")

    with get_db() as conn:
        with conn.cursor() as cur:
            # 获取适配器配置
            cur.execute(
                "SELECT column_mapping, date_format, skip_rows FROM pdca.broker_adapter WHERE broker_name = %s AND is_active = TRUE",
                (broker_name,),
            )
            adapter = cur.fetchone()
            if not adapter:
                raise HTTPException(status_code=400, detail=PDCAError.BROKER_NOT_SUPPORTED.detail())

            column_mapping, date_format, skip_rows = adapter

    # 解析 Excel
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail=PDCAError.EXCEL_PARSE_FAILED.detail())

    # 跳过表头行
    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[skip_rows:]

    if not data_rows:
        raise HTTPException(status_code=400, detail=PDCAError.EXCEL_EMPTY.detail())

    # 获取表头映射
    header_row = rows[skip_rows - 1] if skip_rows > 0 else data_rows[0]
    if skip_rows == 0:
        data_rows = data_rows[1:]

    # 构建列名 → 索引映射
    col_index = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            col_index[str(cell).strip()] = idx

    # 解析数据
    parsed_records = []
    errors = []
    for row_idx, row in enumerate(data_rows, skip_rows + 2):
        if all(cell is None for cell in row):
            continue
        try:
            record = {}
            for db_field, excel_col in column_mapping.items():
                if excel_col in col_index:
                    value = row[col_index[excel_col]]
                    record[db_field] = value
                else:
                    record[db_field] = None
            parsed_records.append(record)
        except Exception as e:
            errors.append({"row": row_idx, "reason": str(e)})

    return ApiResponse(
        code=200,
        message="success",
        data={
            "total": len(parsed_records) + len(errors),
            "success": len(parsed_records),
            "failed": len(errors),
            "errors": errors,
            "records": parsed_records,
        },
    )


@router.post("/import/confirm", response_model=ApiResponse)
async def confirm_import(records: list[dict]):
    """确认导入解析后的数据"""
    if not records:
        raise HTTPException(status_code=400, detail=PDCAError.IMPORT_EMPTY.detail())

    with get_db() as conn:
        with conn.cursor() as cur:
            imported_ids = []
            errors = []
            for idx, record in enumerate(records):
                try:
                    cur.execute(
                        """
                        INSERT INTO pdca.trading_record
                            (pdca_cycle_id, code, security_name, long_short, entry_date, entry_price,
                             quantity, commission_entry, settlement_currency)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                        """,
                        (
                            record.get("pdca_cycle_id", 1),
                            record.get("code"),
                            record.get("security_name"),
                            record.get("long_short", "long"),
                            record.get("entry_date"),
                            float(record.get("entry_price", 0)),
                            int(record.get("quantity", 0)),
                            float(record.get("commission_entry", 0)),
                            record.get("settlement_currency", "CNY"),
                        ),
                    )
                    imported_ids.append(cur.fetchone()[0])
                except Exception as e:
                    errors.append({"row": idx + 1, "reason": str(e)})
                    conn.rollback()
                    # 重新开始事务
                    conn.cursor()

            conn.commit()

    return ApiResponse(
        code=200,
        message="success",
        data={
            "imported": len(imported_ids),
            "failed": len(errors),
            "errors": errors,
            "ids": imported_ids,
        },
    )